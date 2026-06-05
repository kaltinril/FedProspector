"""Tests for targeted bug fixes in attachment_text_extractor.

Covers three recurring text-extraction failures:
  Fix 1: _extract_xlsx must skip chart sheets (no iter_rows) without crashing.
  Fix 2: _check_ole2_encryption raises _UnsupportedType (terminal) for
         IRM/DRM-protected OLE2 files, and _extract_file maps that to
         status "unsupported".
  Fix 3: _extract_file(None, ...) returns status "failed" with a no-file_path
         message and does not raise.

No real database or live files are touched (per project memory).
"""

import os
import sys

import pytest

# Ensure fed_prospector/ is on sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from etl import attachment_text_extractor as ate
from etl.attachment_text_extractor import (
    _check_ole2_encryption,
    _extract_file,
    _extract_xlsx,
    _UnsupportedType,
)


# ---------------------------------------------------------------------------
# Fix 1: _extract_xlsx skips Chartsheet (no iter_rows) without crashing.
# ---------------------------------------------------------------------------

class _FakeWorksheet:
    """Minimal worksheet exposing iter_rows."""

    def __init__(self, rows):
        self._rows = rows

    def iter_rows(self, values_only=True):
        return iter(self._rows)


class _FakeChartsheet:
    """A chart sheet has no iter_rows attribute (like openpyxl Chartsheet)."""


class _FakeWorkbook:
    def __init__(self, sheets):
        # sheets: dict name -> sheet object; sheetnames preserves order
        self._sheets = sheets

    @property
    def sheetnames(self):
        return list(self._sheets.keys())

    def __getitem__(self, name):
        return self._sheets[name]

    def close(self):
        pass


def test_extract_xlsx_skips_chartsheet(monkeypatch):
    """A workbook containing a chart sheet extracts the data sheet and does
    not raise AttributeError on the chart sheet."""
    data_rows = [("Name", "Value"), ("alpha", 1), ("beta", 2)]
    fake_wb = _FakeWorkbook(
        {
            "Chart1": _FakeChartsheet(),  # no iter_rows -> must be skipped
            "Data": _FakeWorksheet(data_rows),
        }
    )

    monkeypatch.setattr(
        ate, "load_workbook", lambda *a, **k: fake_wb, raising=False
    )
    # _extract_xlsx imports load_workbook from openpyxl inside the function,
    # so patch at the openpyxl source too.
    import openpyxl
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *a, **k: fake_wb)

    text, page_count, is_scanned = _extract_xlsx("ignored.xlsx")

    assert "## Data" in text
    assert "alpha" in text
    assert "beta" in text
    assert "Chart1" not in text  # chart sheet produced no output
    assert page_count == 0
    assert is_scanned is False


# ---------------------------------------------------------------------------
# Fix 2: _check_ole2_encryption raises _UnsupportedType for IRM/DRM files.
# ---------------------------------------------------------------------------

class _FakeOle:
    def __init__(self, streams):
        self._streams = streams

    def listdir(self):
        # olefile returns lists of path components; join with "/" in caller
        return [s.split("/") for s in self._streams]

    def close(self):
        pass


@pytest.fixture
def ole2_header_file(tmp_path):
    """A tiny file beginning with the OLE2 magic signature."""
    p = tmp_path / "fake.doc"
    p.write_bytes(b"\xd0\xcf\x11\xe0" + b"\x00" * 16)
    return str(p)


def _patch_olefile(monkeypatch, streams):
    import olefile

    monkeypatch.setattr(olefile, "isOleFile", lambda *a, **k: True)
    monkeypatch.setattr(
        olefile, "OleFileIO", lambda *a, **k: _FakeOle(streams)
    )


def test_check_ole2_encryption_raises_unsupported_for_irm(
    monkeypatch, ole2_header_file
):
    """An EncryptedPackage stream => terminal _UnsupportedType, not RuntimeError."""
    _patch_olefile(monkeypatch, ["EncryptedPackage"])

    with pytest.raises(_UnsupportedType):
        _check_ole2_encryption(ole2_header_file)


def test_check_ole2_encryption_not_runtimeerror(monkeypatch, ole2_header_file):
    """Confirm the raised type is specifically _UnsupportedType (subclass of
    Exception) and that a bare RuntimeError catch would NOT have caught it."""
    _patch_olefile(monkeypatch, ["\x06DataSpaces"])

    try:
        _check_ole2_encryption(ole2_header_file)
        pytest.fail("expected _UnsupportedType to be raised")
    except _UnsupportedType:
        pass  # correct
    except RuntimeError:
        pytest.fail("raised RuntimeError instead of _UnsupportedType")


def test_check_ole2_encryption_clean_file_ok(monkeypatch, ole2_header_file):
    """A normal OLE2 file with no encryption streams does not raise."""
    _patch_olefile(monkeypatch, ["WordDocument", "1Table"])

    # Should return None without raising.
    assert _check_ole2_encryption(ole2_header_file) is None


def test_extract_file_unsupported_when_handler_raises_unsupported(
    monkeypatch, tmp_path
):
    """_extract_file maps a handler raising _UnsupportedType to status
    'unsupported' (terminal)."""
    f = tmp_path / "doc.docx"
    f.write_bytes(b"\xd0\xcf\x11\xe0" + b"\x00" * 16)

    def _raising_handler(_path):
        raise _UnsupportedType("IRM-protected")

    # Force handler resolution to our raising handler and disable magic fallback.
    monkeypatch.setattr(ate, "_resolve_handler", lambda *a, **k: _raising_handler)
    monkeypatch.setattr(ate, "_detect_type_by_magic", lambda *a, **k: None)

    result = _extract_file(
        file_path="doc.docx",
        content_type="application/msword",
        ext=".docx",
        filename="doc.docx",
        attachment_dir=str(tmp_path),
    )

    assert result["status"] == "unsupported"


# ---------------------------------------------------------------------------
# Fix 3: _extract_file(None, ...) returns failed with a no-file_path message.
# ---------------------------------------------------------------------------

def test_extract_file_none_path_returns_failed():
    """A NULL file_path returns status 'failed' (re-triggerable) and does not
    raise a TypeError."""
    result = _extract_file(
        file_path=None,
        content_type="application/pdf",
        ext=".pdf",
        filename="x.pdf",
        attachment_dir="E:/fedprospector/attachments",
    )

    assert result["status"] == "failed"
    assert "file_path" in result["error"]


def test_extract_file_empty_path_returns_failed():
    """An empty-string file_path is also treated as unavailable."""
    result = _extract_file(
        file_path="",
        content_type="application/pdf",
        ext=".pdf",
        filename="x.pdf",
        attachment_dir="E:/fedprospector/attachments",
    )

    assert result["status"] == "failed"
    assert "file_path" in result["error"]

"""Generate FedProspect-Schema-Reference.xlsx from DDL files."""

import re
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Column descriptions ──────────────────────────────────────────────────
COLUMN_DESCRIPTIONS = {
    # Common ETL columns
    "record_hash": "SHA-256 hash for change detection",
    "first_loaded_at": "Timestamp of initial ETL load",
    "last_loaded_at": "Timestamp of most recent ETL load",
    "last_load_id": "FK to etl_load_log for traceability",
    "created_at": "Row creation timestamp",
    "updated_at": "Row last-update timestamp",
    "raw_json": "Full API response JSON for replay/rebuild",
    "raw_record_hash": "SHA-256 hash of raw JSON payload",
    "processed": "Y/N staging processing flag",
    "processed_at": "When staging row was processed",
    "error_message": "Error details if processing failed",
    "load_id": "FK to etl_load_log",

    # Entity
    "uei_sam": "SAM.gov Unique Entity Identifier (12 chars)",
    "uei_duns": "Legacy DUNS number (9 digits)",
    "cage_code": "Commercial and Government Entity code",
    "dodaac": "DoD Activity Address Code",
    "registration_status": "SAM.gov registration status (A=Active)",
    "legal_business_name": "Official registered business name",
    "dba_name": "Doing-business-as name",
    "primary_naics": "Primary NAICS code for this entity",
    "entity_url": "Entity website URL",
    "entity_structure_code": "Legal structure code (LLC, Corp, etc.)",
    "exclusion_status_flag": "Whether entity has active exclusions",
    "record_hash": "SHA-256 hash for change detection",

    # Opportunity
    "notice_id": "SAM.gov opportunity identifier",
    "title": "Opportunity title",
    "solicitation_number": "Government solicitation/RFP number",
    "department_name": "Issuing federal department",
    "sub_tier": "Issuing sub-tier agency",
    "office": "Issuing contracting office",
    "posted_date": "Date opportunity was posted",
    "response_deadline": "Proposal submission deadline",
    "archive_date": "Date opportunity was archived",
    "type": "Opportunity type (Solicitation, Award, etc.)",
    "base_type": "Base opportunity type before modifications",
    "set_aside_code": "Set-aside type code (SBA, WOSB, 8A, etc.)",
    "set_aside_description": "Human-readable set-aside description",
    "classification_code": "Product/Service classification code",
    "naics_code": "NAICS industry classification code",
    "pop_state": "Place of performance state (ISO 3166-2)",
    "pop_zip": "Place of performance ZIP code",
    "pop_country": "Place of performance country code",
    "pop_city": "Place of performance city",
    "active": "Y/N whether opportunity is active",
    "award_number": "Contract award number if awarded",
    "award_date": "Date contract was awarded",
    "award_amount": "Dollar amount of award",
    "awardee_uei": "UEI of winning contractor",
    "awardee_name": "Name of winning contractor",
    "awardee_cage_code": "CAGE code of winning contractor",
    "description_url": "URL to fetch description via SAM.gov API",
    "description_text": "Cached description fetched from description_url",
    "link": "Direct link to opportunity on SAM.gov",
    "resource_links": "JSON array of additional resource URLs",
    "contracting_office_id": "Contracting office identifier",
    "full_parent_path_name": "Full agency hierarchy path (names)",
    "full_parent_path_code": "Full agency hierarchy path (codes)",

    # Attachment
    "attachment_id": "Auto-increment attachment identifier",
    "url": "Download URL for the attachment",
    "filename": "Original filename of attachment",
    "content_type": "MIME type of file content",
    "file_size_bytes": "File size in bytes",
    "file_path": "Local filesystem path to downloaded file",
    "extracted_text": "Full text extracted from document",
    "page_count": "Number of pages in document",
    "is_scanned": "Whether document is a scanned image (OCR needed)",
    "ocr_quality": "Quality of OCR extraction (good/fair/poor)",
    "download_status": "File download state machine status",
    "extraction_status": "Text extraction state machine status",
    "content_hash": "SHA-256 hash of file content",
    "text_hash": "SHA-256 hash of extracted text",
    "downloaded_at": "When file was downloaded",
    "extracted_at": "When text was extracted",

    # Attachment intel
    "intel_id": "Auto-increment intel record identifier",
    "extraction_method": "How intel was extracted (keyword/heuristic/AI)",
    "source_text_hash": "Hash of source text used for extraction",
    "clearance_required": "Y/N security clearance needed",
    "clearance_level": "Required clearance level (Secret, TS, etc.)",
    "clearance_scope": "Scope of clearance requirement",
    "clearance_details": "Free-text clearance details",
    "eval_method": "Evaluation methodology (LPTA, best value, etc.)",
    "eval_details": "Evaluation criteria details",
    "vehicle_type": "Contract vehicle type (GSA, BPA, IDIQ, etc.)",
    "vehicle_details": "Contract vehicle details",
    "is_recompete": "Y/N whether this is a recompete",
    "incumbent_name": "Current incumbent contractor name",
    "recompete_details": "Recompete analysis details",
    "scope_summary": "AI-generated scope summary",
    "period_of_performance": "Contract period of performance",
    "labor_categories": "JSON array of required labor categories",
    "key_requirements": "JSON array of key requirements",
    "overall_confidence": "Confidence level of extraction (high/medium/low)",
    "confidence_details": "JSON breakdown of confidence by field",

    # Intel source
    "source_id": "Auto-increment source reference ID",
    "field_name": "Which intel field this source supports",
    "source_filename": "Filename where text was found",
    "page_number": "Page number in source document",
    "char_offset_start": "Character offset start in extracted text",
    "char_offset_end": "Character offset end in extracted text",
    "matched_text": "Exact text that matched the pattern",
    "surrounding_context": "Text context around the match",
    "pattern_name": "Name of regex/heuristic pattern used",
    "confidence": "Confidence level for this source match",

    # Federal hierarchy
    "fh_org_id": "Federal hierarchy organization ID",
    "fh_org_name": "Federal organization name",
    "fh_org_type": "Organization type (department, agency, office)",
    "status": "Organization status",
    "agency_code": "Agency code identifier",
    "oldfpds_office_code": "Legacy FPDS office code",
    "cgac": "Common Government-wide Accounting Classification",
    "parent_org_id": "Parent organization in hierarchy",
    "level": "Depth level in hierarchy tree",

    # FPDS contract
    "contract_id": "FPDS contract identifier (PIID)",
    "idv_piid": "Indefinite Delivery Vehicle PIID",
    "modification_number": "Contract modification sequence number",
    "transaction_number": "Transaction sequence number",
    "agency_id": "Contracting agency code",
    "agency_name": "Contracting agency name",
    "contracting_office_name": "Contracting office name",
    "funding_agency_id": "Funding agency code",
    "funding_agency_name": "Funding agency name",
    "vendor_uei": "Contractor UEI",
    "vendor_name": "Contractor name",
    "vendor_duns": "Legacy contractor DUNS",
    "date_signed": "Contract signature date",
    "effective_date": "Contract effective date",
    "completion_date": "Contract completion date",
    "last_modified_date": "Last modification date",
    "dollars_obligated": "Dollars obligated on contract",
    "base_and_all_options": "Total value including all options",
    "psc_code": "Product/Service Code",
    "set_aside_type": "Set-aside type for this contract",
    "type_of_contract": "Contract type code",
    "description": "Description text",
    "extent_competed": "Competition extent code",
    "number_of_offers": "Number of offers received",
    "solicitation_number": "Solicitation number linked to contract",
    "ultimate_completion_date": "Ultimate contract end date including options",
    "co_bus_size_determination": "Business size determination",

    # GSA labor rate
    "labor_category": "Labor category title",
    "education_level": "Required education level",
    "min_years_experience": "Minimum years of experience required",
    "current_price": "Current hourly rate",
    "next_year_price": "Next option year hourly rate",
    "second_year_price": "Second option year hourly rate",
    "schedule": "GSA schedule name",
    "contractor_name": "GSA schedule holder name",
    "sin": "Special Item Number(s)",
    "business_size": "Business size (S=Small, O=Other)",
    "security_clearance": "Required clearance level",
    "worksite": "Work location requirements",

    # SAM exclusion
    "uei": "Entity UEI (may differ from entity.uei_sam)",
    "entity_name": "Excluded entity name",
    "exclusion_type": "Type of exclusion (Ineligible, Prohibited, etc.)",
    "exclusion_program": "Exclusion program (Reciprocal, NonProcurement, etc.)",
    "excluding_agency_code": "Agency that imposed exclusion",
    "excluding_agency_name": "Name of excluding agency",
    "activation_date": "Exclusion start date",
    "termination_date": "Exclusion end date",
    "is_fascsa_order": "FASCSA order indicator",
    "classification_type": "Exclusion classification",
    "termination_type": "How exclusion terminates",

    # Subaward
    "prime_piid": "Prime contract PIID",
    "prime_agency_id": "Prime contract agency code",
    "prime_agency_name": "Prime contract agency name",
    "prime_uei": "Prime contractor UEI",
    "prime_name": "Prime contractor name",
    "sub_uei": "Subcontractor UEI",
    "sub_name": "Subcontractor name",
    "sub_amount": "Subaward dollar amount",
    "sub_date": "Subaward date",
    "sub_description": "Subaward description",
    "sub_business_type": "Subcontractor business type codes",

    # Reference tables
    "code_level": "Depth in NAICS hierarchy (2-6 digit)",
    "level_name": "Human name for hierarchy level",
    "parent_code": "Parent NAICS code in hierarchy",
    "year_version": "NAICS revision year",
    "is_active": "Y/N active status flag",
    "footnote_id": "Reference to NAICS footnote",
    "size_standard": "SBA size standard (revenue or employees)",
    "size_type": "R=Revenue, E=Employees",
    "effective_date": "When size standard took effect",
    "industry_description": "SBA industry description",
    "section": "Footnote section identifier",
    "psc_code": "Product/Service Code",
    "psc_name": "PSC code name/description",
    "start_date": "PSC code effective start date",
    "end_date": "PSC code end date (null if current)",
    "full_description": "Full PSC description text",
    "psc_includes": "What this PSC code includes",
    "psc_excludes": "What this PSC code excludes",
    "psc_notes": "Additional PSC notes",
    "parent_psc_code": "Parent PSC in hierarchy",
    "category_type": "P=Product, S=Service, R=R&D",
    "level1_category_code": "Top-level category code",
    "level1_category": "Top-level category name",
    "level2_category_code": "Second-level category code",
    "level2_category": "Second-level category name",
    "country_name": "Country full name",
    "two_code": "ISO 3166-1 alpha-2 code",
    "three_code": "ISO 3166-1 alpha-3 code",
    "numeric_code": "ISO 3166-1 numeric code",
    "independent": "Whether country is independent (Yes/No)",
    "is_iso_standard": "Whether code is ISO standard",
    "sam_gov_recognized": "Whether recognized by SAM.gov",
    "state_code": "US state/territory code",
    "state_name": "US state/territory name",
    "country_code": "Associated country code",
    "fips_code": "FIPS county code",
    "county_name": "County name",
    "business_type_code": "SAM.gov business type code",
    "classification": "Business type classification",
    "category": "Business type category",
    "is_socioeconomic": "Y/N socioeconomic indicator",
    "is_small_business_related": "Y/N small business indicator",
    "structure_code": "Entity legal structure code",
    "set_aside_code": "Set-aside type code",
    "is_small_business": "Y/N small business set-aside",
    "sba_type_code": "SBA certification type code",
    "program_name": "SBA program name",

    # ETL tables
    "source_system": "Data source name (SAM, USASpending, etc.)",
    "load_type": "Type of load (FULL, INCREMENTAL, etc.)",
    "started_at": "Load start timestamp",
    "completed_at": "Load completion timestamp",
    "records_read": "Total records read from source",
    "records_inserted": "New records inserted",
    "records_updated": "Existing records updated",
    "records_unchanged": "Records with no changes (hash match)",
    "records_errored": "Records that failed processing",
    "parameters": "JSON load parameters (date range, filters)",
    "source_file": "Source file path for file-based loads",
    "record_identifier": "Identifier of the errored record",
    "error_type": "Category of error",
    "raw_data": "Raw data that caused the error",
    "rule_id": "Auto-increment rule identifier",
    "rule_name": "Data quality rule name",
    "target_table": "Table this rule applies to",
    "target_column": "Column this rule applies to",
    "rule_type": "Rule type (REQUIRED, RANGE, ENUM, etc.)",
    "rule_definition": "JSON rule definition/parameters",
    "priority": "Rule execution priority (lower = first)",
    "request_id": "Auto-increment request identifier",
    "request_type": "Type of load request",
    "lookup_key": "Identifier to look up (PIID, UEI, etc.)",
    "lookup_key_type": "Type of lookup key (PIID, UEI, etc.)",
    "requested_by": "User who made the request",
    "requested_at": "When request was created",
    "result_summary": "JSON summary of load results",
    "request_date": "Date of rate limit tracking",
    "requests_made": "Number of API requests made",
    "max_requests": "Maximum allowed requests per day",
    "last_request_at": "Timestamp of last API request",
    "snapshot_id": "Auto-increment snapshot ID",
    "checked_at": "When health check was run",
    "overall_status": "Overall system health status",
    "results_json": "JSON health check results",
    "alert_count": "Number of health alerts",
    "error_count": "Number of health errors",
    "stale_source_count": "Number of stale data sources",

    # AI usage tracking
    "usage_id": "Auto-increment usage log identifier",
    "model": "AI model name (e.g. claude-3-haiku)",
    "input_tokens": "Number of input tokens consumed",
    "output_tokens": "Number of output tokens generated",
    "cache_read_tokens": "Tokens read from prompt cache",
    "cache_write_tokens": "Tokens written to prompt cache",
    "cost_usd": "Estimated cost in USD",

    # Application tables
    "organization_id": "FK to organization (multi-tenant isolation)",
    "user_id": "FK to app_user",
    "name": "Organization display name",
    "slug": "URL-safe organization identifier",
    "max_users": "Maximum users allowed in organization",
    "subscription_tier": "Subscription level (trial, basic, pro)",
    "legal_name": "Organization legal business name",
    "ein": "Employer Identification Number",
    "entity_structure": "Legal entity structure",
    "phone": "Contact phone number",
    "website": "Organization website URL",
    "annual_revenue": "Annual revenue for size standard checks",
    "employee_count": "Number of employees",
    "fiscal_year_end_month": "Month fiscal year ends (1-12)",
    "profile_completed": "Y/N org profile is complete",
    "profile_completed_at": "When profile was completed",
    "username": "Login username (unique)",
    "display_name": "User display name",
    "email": "User email address",
    "password_hash": "Bcrypt password hash",
    "role": "Application role (USER, ADMIN)",
    "last_login_at": "Last successful login timestamp",
    "is_org_admin": "Y/N organization admin flag",
    "mfa_enabled": "Y/N multi-factor auth enabled",
    "org_role": "Organization role (owner, admin, member)",
    "invited_by": "User who sent the invitation",
    "invite_accepted_at": "When invite was accepted",
    "is_system_admin": "Whether user is a system admin",
    "force_password_change": "Y/N must change password on next login",
    "failed_login_attempts": "Consecutive failed login count",
    "locked_until": "Account locked until this timestamp",
    "prospect_id": "FK to prospect",
    "source": "How prospect was created (MANUAL, AUTO, SEARCH)",
    "assigned_to": "User assigned to work this prospect",
    "capture_manager_id": "Capture manager for this opportunity",
    "proposal_status": "Proposal workflow status",
    "decision_date": "Go/no-go decision date",
    "bid_submitted_date": "When bid was submitted",
    "estimated_value": "Estimated contract value",
    "estimated_effort_hours": "Estimated effort in hours",
    "win_probability": "Probability of win (0-100%)",
    "go_no_go_score": "Go/no-go evaluation score",
    "teaming_required": "Y/N teaming arrangement needed",
    "estimated_proposal_cost": "Cost to prepare proposal",
    "estimated_gross_margin_pct": "Expected gross margin percentage",
    "proposal_due_days": "Days until proposal is due",
    "outcome": "Win/loss outcome",
    "outcome_date": "Date of outcome",
    "outcome_notes": "Notes about outcome",
    "contract_award_id": "Linked contract award identifier",
    "note_id": "Auto-increment note identifier",
    "note_type": "Type of note (COMMENT, STATUS_CHANGE, etc.)",
    "note_text": "Note content",
    "search_id": "Auto-increment saved search identifier",
    "search_name": "User-defined search name",
    "filter_criteria": "JSON search filter parameters",
    "notification_enabled": "Y/N email notifications for new matches",
    "last_run_at": "When search was last executed",
    "last_new_results": "Number of new results in last run",
    "auto_prospect_enabled": "Y/N auto-create prospects from matches",
    "min_pwin_score": "Minimum pWin score for auto-prospect",
    "auto_assign_to": "User to auto-assign new prospects",
    "last_auto_run_at": "Last auto-prospect run timestamp",
    "last_auto_created": "Count of prospects auto-created last run",
    "notification_id": "Auto-increment notification identifier",
    "notification_type": "Notification category",
    "message": "Notification message body",
    "entity_type": "Type of related entity",
    "entity_id": "ID of related entity",
    "is_read": "Y/N notification read flag",
    "read_at": "When notification was read",

    # Session
    "session_id": "Auto-increment session identifier",
    "token_hash": "SHA-256 hash of JWT access token",
    "refresh_token_hash": "SHA-256 hash of refresh token",
    "issued_at": "When token was issued",
    "expires_at": "When token expires",
    "revoked_at": "When token was revoked",
    "revoked_reason": "Reason for revocation",
    "ip_address": "Client IP address",
    "user_agent": "Client user agent string",

    # Invite
    "invite_id": "Auto-increment invite identifier",
    "invite_code": "Unique invitation code",
    "accepted_at": "When invitation was accepted",

    # Contracting officer
    "officer_id": "Auto-increment officer identifier",
    "full_name": "Officer full name",
    "fax": "Fax number",
    "officer_type": "Type (Contracting Officer, Specialist, etc.)",

    # Opportunity POC
    "poc_id": "Auto-increment POC identifier",
    "officer_id": "FK to contracting_officer",
    "poc_type": "Contact type (PRIMARY, SECONDARY)",

    # Proposal
    "proposal_id": "Auto-increment proposal identifier",
    "proposal_number": "User-defined proposal number",
    "submission_deadline": "Proposal submission deadline",
    "submitted_at": "When proposal was submitted",
    "win_probability_pct": "Estimated win probability percentage",
    "lessons_learned": "Post-award lessons learned",

    # Proposal document
    "document_id": "Auto-increment document identifier",
    "document_type": "Document category (TECHNICAL, COST, etc.)",
    "file_name": "Original uploaded filename",
    "uploaded_by": "User who uploaded the document",
    "uploaded_at": "Upload timestamp",
    "notes": "Additional notes",

    # Proposal milestone
    "milestone_id": "Auto-increment milestone identifier",
    "milestone_name": "Milestone name/description",
    "due_date": "Milestone due date",
    "completed_date": "When milestone was completed",

    # Activity log
    "activity_id": "Auto-increment activity identifier",
    "action": "Action performed (CREATE, UPDATE, DELETE, etc.)",
    "details": "JSON details of the action",

    # Organization child tables
    "certification_type": "Certification type (WOSB, 8a, HUBZone, etc.)",
    "certifying_agency": "Agency that issued certification",
    "certification_number": "Certification number",
    "expiration_date": "Certification expiration date",
    "size_standard_met": "Y/N organization meets size standard",
    "is_primary": "Y/N primary designation",
    "contract_number": "Past performance contract number",
    "contract_value": "Past performance contract value",
    "period_start": "Contract period start",
    "period_end": "Contract period end",
    "relationship": "Entity relationship type (SELF, COMPETITOR, etc.)",
    "partner_uei": "UEI for JV partnership filings",
    "added_by": "User who added this link",

    # USASpending
    "generated_unique_award_id": "USASpending unique award identifier",
    "piid": "Procurement Instrument Identifier",
    "fain": "Federal Award Identification Number",
    "uri": "Unique Record Identifier",
    "award_type": "Award type (contract, grant, etc.)",
    "award_description": "Award description text",
    "recipient_name": "Recipient/incumbent name",
    "recipient_uei": "Recipient UEI",
    "recipient_parent_name": "Parent organization name",
    "recipient_parent_uei": "Parent organization UEI",
    "total_obligation": "Total dollars obligated",
    "base_and_all_options_value": "Total value including all options",
    "awarding_agency_name": "Awarding agency name",
    "awarding_sub_agency_name": "Awarding sub-agency name",
    "funding_agency_name": "Funding agency name",
    "naics_description": "NAICS code description",
    "type_of_set_aside": "Set-aside type code",
    "type_of_set_aside_description": "Set-aside description",
    "solicitation_identifier": "Linked solicitation number",
    "fiscal_year": "Federal fiscal year",
    "fpds_enriched_at": "When FPDS enrichment was applied",
    "deleted_at": "Soft-delete timestamp (preserves FK refs)",
    "award_id": "FK to usaspending_award",
    "action_date": "Transaction action date",
    "action_type": "Transaction action type code",
    "action_type_description": "Transaction action type description",
    "federal_action_obligation": "Transaction obligation amount",

    # Checkpoint
    "checkpoint_id": "Auto-increment checkpoint ID",
    "csv_file_name": "Name of CSV file being loaded",
    "completed_batches": "Number of batches completed",
    "total_rows_loaded": "Total rows loaded so far",
    "archive_hash": "SHA-256 hash for FY dedup",

    # Team member
    "app_user_id": "FK to app_user for internal team members",
    "proposed_hourly_rate": "Proposed billing rate",
    "commitment_pct": "Percentage of time committed",

    # Entity sub-tables
    "address_type": "Address type (physical, mailing)",
    "address_line_1": "Address line 1",
    "address_line_2": "Address line 2",
    "city": "City name",
    "state_or_province": "State or province",
    "zip_code": "ZIP/postal code",
    "zip_code_plus4": "ZIP+4 extension",
    "congressional_district": "Congressional district",
    "sba_small_business": "SBA small business designation",
    "naics_exception": "NAICS exception string",
    "sba_type_desc": "SBA certification type description",
    "certification_entry_date": "When certification was granted",
    "certification_exit_date": "When certification expired/ended",
    "first_name": "Contact first name",
    "middle_initial": "Contact middle initial",
    "last_name": "Contact last name",
    "msa_code": "Metropolitan Statistical Area code",
    "msa_name": "Metropolitan Statistical Area name",
    "field_name": "Changed field name",
    "old_value": "Previous field value",
    "new_value": "New field value",
    "changed_at": "When change was detected",

    # History
    "parent_notice_id": "Parent opportunity notice_id",
    "child_notice_id": "Child opportunity notice_id",
    "relationship_type": "Relationship type (RFI_TO_RFP, etc.)",
    "created_by": "User who created the relationship",
}


def parse_create_tables(sql_text):
    """Parse CREATE TABLE statements from SQL text."""
    tables = {}
    # Find all CREATE TABLE blocks
    pattern = re.compile(
        r'CREATE\s+TABLE\s+IF\s+NOT\s+EXISTS\s+(\w+)\s*\((.*?)\)\s*ENGINE',
        re.DOTALL | re.IGNORECASE
    )
    for match in pattern.finditer(sql_text):
        table_name = match.group(1)
        body = match.group(2)
        columns = parse_columns(body, table_name)
        tables[table_name] = columns
    return tables


def parse_columns(body, table_name):
    """Parse column definitions from CREATE TABLE body."""
    columns = []
    pk_cols = set()
    indexed_cols = set()

    # Find PRIMARY KEY
    pk_match = re.search(r'PRIMARY\s+KEY\s*\(([^)]+)\)', body, re.IGNORECASE)
    if pk_match:
        for col in pk_match.group(1).split(','):
            pk_cols.add(col.strip().strip('`'))

    # Find indexes
    for idx_match in re.finditer(r'(?:INDEX|KEY|UNIQUE\s+(?:KEY|INDEX))\s+\w+\s*\(([^)]+)\)', body, re.IGNORECASE):
        for col in idx_match.group(1).split(','):
            col_clean = re.sub(r'\(\d+\)', '', col).strip().strip('`')
            indexed_cols.add(col_clean)

    # Parse each line
    lines = body.split('\n')
    for line in lines:
        line = line.strip().rstrip(',')
        if not line:
            continue

        # Skip constraints, indexes, keys
        if re.match(r'^\s*(PRIMARY\s+KEY|INDEX|KEY|UNIQUE|CONSTRAINT|FOREIGN|FULLTEXT)', line, re.IGNORECASE):
            continue

        # Column definition pattern
        col_match = re.match(
            r'^(\w+)\s+([\w()]+(?:\s*\([^)]*\))?(?:\s+UNSIGNED)?)',
            line, re.IGNORECASE
        )
        if not col_match:
            continue

        col_name = col_match.group(1)
        # Skip SQL keywords that aren't columns
        if col_name.upper() in ('PRIMARY', 'INDEX', 'KEY', 'UNIQUE', 'CONSTRAINT', 'FOREIGN', 'FULLTEXT', 'CHECK'):
            continue

        data_type = col_match.group(2).upper()
        # Clean up data type
        data_type = re.sub(r'\s+', ' ', data_type).strip()

        nullable = 'NOT NULL' not in line.upper()
        # Check for AUTO_INCREMENT PRIMARY KEY inline
        is_pk = col_name in pk_cols or 'AUTO_INCREMENT PRIMARY KEY' in line.upper() or 'PRIMARY KEY' in line.upper()
        if is_pk and 'AUTO_INCREMENT' in line.upper():
            pk_cols.add(col_name)

        is_indexed = col_name in indexed_cols or is_pk

        # Default value
        default = None
        def_match = re.search(r"DEFAULT\s+'([^']*)'", line, re.IGNORECASE)
        if not def_match:
            def_match = re.search(r'DEFAULT\s+(CURRENT_TIMESTAMP|NULL|\d+(?:\.\d+)?)', line, re.IGNORECASE)
        if def_match:
            default = def_match.group(1)

        # Comment
        comment_match = re.search(r"COMMENT\s+'([^']*)'", line, re.IGNORECASE)
        desc = comment_match.group(1) if comment_match else COLUMN_DESCRIPTIONS.get(col_name, '')

        # Handle inline comment from SQL
        inline = re.search(r'--\s*(.+)$', line)
        if inline and not desc:
            desc = inline.group(1).strip()
        elif inline and desc:
            # Prefer the inline comment if it's more specific
            pass

        columns.append({
            'name': col_name,
            'data_type': data_type,
            'nullable': 'YES' if nullable else 'NO',
            'default': default or '',
            'pk': 'PK' if is_pk else '',
            'indexed': 'Y' if is_indexed else '',
            'description': desc,
        })

    return columns


def create_workbook():
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    table_name_font = Font(bold=True, size=11)
    table_name_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'),
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Read all DDL files
    ddl_dir = r'C:\git\fedProspect\fed_prospector\db\schema\tables'
    all_tables = {}
    for f in sorted(os.listdir(ddl_dir)):
        if f.endswith('.sql'):
            with open(os.path.join(ddl_dir, f), 'r', encoding='utf-8') as fh:
                sql = fh.read()
            all_tables.update(parse_create_tables(sql))

    # Sheet definitions: (sheet_name, [(table_name, display_label), ...])
    sheets = [
        ("Opportunity", [
            ("opportunity", "opportunity"),
            ("opportunity_history", "opportunity_history"),
            ("opportunity_relationship", "opportunity_relationship"),
            ("opportunity_attachment", "opportunity_attachment"),
            ("opportunity_attachment_intel", "opportunity_attachment_intel"),
            ("opportunity_intel_source", "opportunity_intel_source"),
            ("opportunity_poc", "opportunity_poc"),
            ("contracting_officer", "contracting_officer"),
        ]),
        ("Award", [
            ("fpds_contract", "fpds_contract"),
            ("federal_organization", "federal_organization"),
            ("usaspending_award", "usaspending_award"),
            ("usaspending_transaction", "usaspending_transaction"),
            ("usaspending_load_checkpoint", "usaspending_load_checkpoint"),
            ("sam_subaward", "sam_subaward"),
            ("gsa_labor_rate", "gsa_labor_rate"),
        ]),
        ("Entity", [
            ("entity", "entity"),
            ("entity_address", "entity_address"),
            ("entity_naics", "entity_naics"),
            ("entity_psc", "entity_psc"),
            ("entity_business_type", "entity_business_type"),
            ("entity_sba_certification", "entity_sba_certification"),
            ("entity_poc", "entity_poc"),
            ("entity_disaster_response", "entity_disaster_response"),
            ("entity_history", "entity_history"),
            ("sam_exclusion", "sam_exclusion"),
        ]),
        ("Reference", [
            ("ref_naics_code", "ref_naics_code"),
            ("ref_sba_size_standard", "ref_sba_size_standard"),
            ("ref_naics_footnote", "ref_naics_footnote"),
            ("ref_psc_code", "ref_psc_code"),
            ("ref_country_code", "ref_country_code"),
            ("ref_state_code", "ref_state_code"),
            ("ref_fips_county", "ref_fips_county"),
            ("ref_business_type", "ref_business_type"),
            ("ref_entity_structure", "ref_entity_structure"),
            ("ref_set_aside_type", "ref_set_aside_type"),
            ("ref_sba_type", "ref_sba_type"),
        ]),
        ("ETL", [
            ("etl_load_log", "etl_load_log"),
            ("etl_load_error", "etl_load_error"),
            ("etl_data_quality_rule", "etl_data_quality_rule"),
            ("etl_rate_limit", "etl_rate_limit"),
            ("etl_health_snapshot", "etl_health_snapshot"),
            ("data_load_request", "data_load_request"),
            ("ai_usage_log", "ai_usage_log"),
        ]),
        ("Application", [
            ("organization", "organization"),
            ("app_user", "app_user"),
            ("app_session", "app_session"),
            ("prospect", "prospect"),
            ("prospect_note", "prospect_note"),
            ("prospect_team_member", "prospect_team_member"),
            ("saved_search", "saved_search"),
            ("notification", "notification"),
            ("proposal", "proposal"),
            ("proposal_document", "proposal_document"),
            ("proposal_milestone", "proposal_milestone"),
            ("activity_log", "activity_log"),
            ("organization_invite", "organization_invite"),
            ("organization_certification", "organization_certification"),
            ("organization_naics", "organization_naics"),
            ("organization_past_performance", "organization_past_performance"),
            ("organization_entity", "organization_entity"),
        ]),
        ("Staging", [
            ("stg_entity_raw", "stg_entity_raw"),
            ("stg_opportunity_raw", "stg_opportunity_raw"),
            ("stg_fpds_award_raw", "stg_fpds_award_raw"),
            ("stg_usaspending_raw", "stg_usaspending_raw"),
            ("stg_exclusion_raw", "stg_exclusion_raw"),
            ("stg_fedhier_raw", "stg_fedhier_raw"),
            ("stg_subaward_raw", "stg_subaward_raw"),
        ]),
    ]

    col_headers = ["Column Name", "Data Type", "Nullable", "Default", "PK", "Index", "Description"]

    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, table_list in sheets:
        ws = wb.create_sheet(title=sheet_name)
        row = 1

        for table_name, display_label in table_list:
            if table_name not in all_tables:
                continue

            columns = all_tables[table_name]

            # Table name header row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(col_headers))
            cell = ws.cell(row=row, column=1, value=display_label)
            cell.font = table_name_font
            cell.fill = table_name_fill
            cell.alignment = Alignment(horizontal='left')
            for c in range(1, len(col_headers) + 1):
                ws.cell(row=row, column=c).fill = table_name_fill
                ws.cell(row=row, column=c).border = thin_border
            row += 1

            # Column headers
            for ci, h in enumerate(col_headers, 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row += 1

            # Data rows
            for ri, col_def in enumerate(columns):
                fill = alt_fill if ri % 2 == 1 else PatternFill()
                values = [
                    col_def['name'],
                    col_def['data_type'],
                    col_def['nullable'],
                    col_def['default'],
                    col_def['pk'],
                    col_def['indexed'],
                    col_def['description'],
                ]
                for ci, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=ci, value=v)
                    cell.border = thin_border
                    cell.fill = fill
                    cell.font = Font(size=10)
                row += 1

            # Blank row between tables
            row += 1

        # Auto-width columns
        for ci in range(1, len(col_headers) + 1):
            max_len = len(col_headers[ci - 1])
            for r in range(1, row):
                val = ws.cell(row=r, column=ci).value
                if val:
                    max_len = max(max_len, len(str(val)))
            # Cap at 60
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 60)

        # Freeze panes (top row of first table header area)
        ws.freeze_panes = "A2"

    # ── Views sheet ──
    views_info = [
        ("v_entity_search", "Entity search with address, business types, certifications, and NAICS details", "entity, entity_address, entity_naics, entity_business_type, entity_sba_certification, ref_naics_code, ref_business_type, ref_entity_structure"),
        ("v_opportunity_latest", "Deduplicated opportunities: one row per solicitation (latest biddable notice)", "opportunity"),
        ("v_target_opportunities", "WOSB/8(a) target opportunities with prospect status overlay", "v_opportunity_latest, prospect, ref_naics_code, ref_psc_code"),
        ("v_competitor_analysis", "Competitor entity profiles with business types, certs, and past performance", "entity, entity_business_type, ref_business_type, ref_entity_structure, ref_naics_code, entity_sba_certification, fpds_contract"),
        ("v_procurement_intelligence", "Opportunity + award history + USASpending + incumbent analysis", "opportunity, fpds_contract, usaspending_award"),
        ("v_incumbent_profile", "Entity profile with SBA certifications and FPDS award history", "entity, entity_sba_certification, fpds_contract"),
        ("v_expiring_contracts", "Contracts expiring within 18 months with incumbent health signals and burn rate", "fpds_contract, entity, entity_sba_certification, usaspending_award, usaspending_transaction"),
        ("v_set_aside_shift", "Set-aside shift analysis: current opportunity vs predecessor contract", "opportunity, fpds_contract"),
        ("v_set_aside_trend", "NAICS-level set-aside trend aggregation by fiscal year", "fpds_contract"),
        ("v_monthly_spend", "Monthly spend breakdown by award for burn rate analysis", "usaspending_transaction"),
        ("v_vendor_market_share", "Vendor market share by NAICS code from FPDS base awards", "fpds_contract"),
        ("ref_psc_code_latest", "Deduplicated PSC codes: latest version of each code by start_date", "ref_psc_code"),
    ]

    ws = wb.create_sheet(title="Views")
    view_headers = ["View Name", "Purpose", "Source Tables"]
    for ci, h in enumerate(view_headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for ri, (vname, vpurpose, vsources) in enumerate(views_info, 2):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, val in enumerate([vname, vpurpose, vsources], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.fill = fill
            cell.font = Font(size=10)

    for ci in range(1, 4):
        max_len = len(view_headers[ci - 1])
        for r in range(1, len(views_info) + 2):
            val = ws.cell(row=r, column=ci).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 80)

    ws.freeze_panes = "A2"

    # ── Summary / cover sheet ──
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.cell(row=1, column=1, value="FedProspect Schema Reference").font = Font(bold=True, size=16, color="1F4E79")
    ws_summary.cell(row=2, column=1, value=f"Generated: {date.today().isoformat()}").font = Font(size=11, italic=True, color="666666")
    ws_summary.cell(row=3, column=1, value="Database: fed_contracts (MySQL 8.4)").font = Font(size=11, color="666666")
    ws_summary.cell(row=5, column=1, value="Sheet").font = Font(bold=True, size=11)
    ws_summary.cell(row=5, column=2, value="Tables").font = Font(bold=True, size=11)
    ws_summary.cell(row=5, column=3, value="Description").font = Font(bold=True, size=11)
    for ci in range(1, 4):
        ws_summary.cell(row=5, column=ci).fill = header_fill
        ws_summary.cell(row=5, column=ci).font = header_font
        ws_summary.cell(row=5, column=ci).border = thin_border

    summary_data = [
        ("Opportunity", "8 tables", "Contract opportunities from SAM.gov, attachments, intel extraction, POCs"),
        ("Award", "6 tables", "FPDS contracts, USASpending awards/transactions, subawards, GSA labor rates"),
        ("Entity", "10 tables", "SAM.gov entity registrations, addresses, NAICS, business types, certs, exclusions"),
        ("Reference", "11 tables", "NAICS, PSC, set-aside, SBA types, FIPS, country/state codes"),
        ("ETL", "6 tables", "Load tracking, error logging, data quality rules, rate limits, health checks"),
        ("Application", "17 tables", "Users, orgs, prospects, proposals, saved searches, notifications, audit log"),
        ("Staging", "7 tables", "Raw API response staging for replay/rebuild capability"),
        ("Views", "12 views", "Pre-built analytical views for search, intelligence, and reporting"),
    ]

    for ri, (sname, count, desc) in enumerate(summary_data, 6):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, val in enumerate([sname, count, desc], 1):
            cell = ws_summary.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.fill = fill
            cell.font = Font(size=10)

    total_tables = sum(1 for t in all_tables if not t.startswith('stg_'))
    staging_tables = sum(1 for t in all_tables if t.startswith('stg_'))
    ws_summary.cell(row=ri + 2, column=1, value=f"Total: {len(all_tables)} tables ({total_tables} production + {staging_tables} staging) + 12 views").font = Font(size=11, bold=True, color="1F4E79")

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 75

    # Save
    out_path = r'C:\git\fedProspect\docs\database\FedProspect-Schema-Reference.xlsx'
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Tables parsed: {len(all_tables)}")
    for t in sorted(all_tables):
        print(f"  {t}: {len(all_tables[t])} columns")


if __name__ == '__main__':
    create_workbook()
